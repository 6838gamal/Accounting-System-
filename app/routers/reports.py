"""
مسارات التقارير
"""
from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Request, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from app.templates_config import templates as _shared_templates
from sqlalchemy.orm import Session
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.dependencies import get_db
from app.services.report_service import ReportService

router = APIRouter(prefix="/reports", tags=["reports"])
templates = _shared_templates


def _parse_date(value: Optional[str], fallback: date) -> date:
    """تحليل التاريخ من نص مع احتياطي عند الخطأ."""
    if not value:
        return fallback
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return fallback


@router.get("", response_class=HTMLResponse)
async def reports_index(request: Request):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    return templates.TemplateResponse("reports/index.html", {"request": request})


import logging
_logger = logging.getLogger(__name__)



@router.get("/sales", response_class=HTMLResponse)
async def sales_report(
    request: Request, db: Session = Depends(get_db),
    start_date: Optional[str] = None, end_date: Optional[str] = None,
):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    today = date.today()
    start = _parse_date(start_date, date(today.year, today.month, 1))
    end = _parse_date(end_date, today)
    context: dict = {"request": request}
    try:
        context["report"] = ReportService(db).get_sales_report(start, end)
    except ValueError as exc:
        _logger.warning("تحقق تاريخ المبيعات | %s", exc)
        context["report"] = {
            "start_date": start, "end_date": end,
            "total_invoices": 0, "total_invoiced": 0.0,
            "total_paid": 0.0, "total_pending": 0.0, "invoices": [],
        }
        context["error"] = str(exc)
    return templates.TemplateResponse("reports/sales.html", context)


@router.get("/clients", response_class=HTMLResponse)
async def clients_report(
    request: Request, db: Session = Depends(get_db),
    start_date: Optional[str] = None, end_date: Optional[str] = None,
):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    today = date.today()
    start = _parse_date(start_date, date(today.year, 1, 1))
    end = _parse_date(end_date, today)
    context: dict = {"request": request, "start": start, "end": end}
    try:
        context["data"] = ReportService(db).get_client_report(start, end)
    except ValueError as exc:
        _logger.warning("تحقق تاريخ العملاء | %s", exc)
        context["data"] = []
        context["error"] = str(exc)
    return templates.TemplateResponse("reports/clients.html", context)


@router.get("/expenses", response_class=HTMLResponse)
async def expenses_report(
    request: Request, db: Session = Depends(get_db),
    start_date: Optional[str] = None, end_date: Optional[str] = None,
):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    today = date.today()
    start = _parse_date(start_date, date(today.year, today.month, 1))
    end = _parse_date(end_date, today)
    context: dict = {"request": request}
    try:
        context["report"] = ReportService(db).get_expense_report(start, end)
    except ValueError as exc:
        _logger.warning("تحقق تاريخ المصروفات | %s", exc)
        context["report"] = {
            "start_date": start, "end_date": end,
            "expenses": [], "by_category": {}, "total": 0.0,
        }
        context["error"] = str(exc)
    return templates.TemplateResponse("reports/expenses.html", context)


@router.get("/profit-loss", response_class=HTMLResponse)
async def profit_loss_report(
    request: Request, db: Session = Depends(get_db),
    start_date: Optional[str] = None, end_date: Optional[str] = None,
):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    today = date.today()
    start = _parse_date(start_date, date(today.year, 1, 1))
    end = _parse_date(end_date, today)
    context: dict = {"request": request}
    try:
        context["report"] = ReportService(db).get_profit_loss(start, end)
    except ValueError as exc:
        _logger.warning("تحقق تاريخ الأرباح والخسائر | %s", exc)
        context["report"] = {
            "start_date": start, "end_date": end,
            "revenue": 0.0, "expenses": 0.0, "profit": 0.0,
        }
        context["error"] = str(exc)
    return templates.TemplateResponse("reports/profit_loss.html", context)


@router.get("/sales/excel")
async def sales_excel(
    request: Request, db: Session = Depends(get_db),
    start_date: Optional[str] = None, end_date: Optional[str] = None,
):
    if not request.session.get("user_id"):
        return RedirectResponse(url="/auth/login", status_code=302)
    today = date.today()
    start = _parse_date(start_date, date(today.year, today.month, 1))
    end = _parse_date(end_date, today)
    try:
        report = ReportService(db).get_sales_report(start, end)
    except ValueError as exc:
        from fastapi.responses import HTMLResponse as _HTML
        return _HTML(content=f"<p>خطأ في التاريخ: {exc}</p>", status_code=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير المبيعات"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    headers = ["رقم الفاتورة", "العميل", "تاريخ الإصدار", "الإجمالي", "المدفوع", "المتبقي", "الحالة"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, inv in enumerate(report["invoices"], 2):
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=inv.client.name if inv.client else "")
        ws.cell(row=row, column=3, value=str(inv.issue_date))
        ws.cell(row=row, column=4, value=float(inv.total))
        ws.cell(row=row, column=5, value=float(inv.paid_amount))
        ws.cell(row=row, column=6, value=float(inv.total - inv.paid_amount))
        ws.cell(row=row, column=7, value=inv.status.value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales-report-{start}.xlsx"},
    )
